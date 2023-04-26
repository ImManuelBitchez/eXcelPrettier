import sys,os
import win32com.client

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from functions.ExcelModules import *
from platform import *
from functions.setup import init
projectFolder = Path(os.getcwd())
module_path = Path(f"{projectFolder}\modules\Modulo1.bas")



if(len(sys.argv) <= 1): 
    print("No csv file founded!")
    print("Please insert at least one file")
    os._exit(1)
else:
    try:
        env_info = init()   
        baseFile = sys.argv[1]
        convertToExcel(baseFile,env_info)
    except Exception as e:
        print(e)
    
    fileList = [f"{projectFolder}\\files\\aws.xlsx",f"{projectFolder}\\files\\marketplace.xlsx",f"{projectFolder}\\files\\refund.xlsx",f"{projectFolder}\\files\\tax.xlsx"]
    #initWorkbook()
    #copyToFile(outfile,filename=outfile)
    #createFile(outfile,fileList)
  