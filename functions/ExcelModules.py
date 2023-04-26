from openpyxl import Workbook,load_workbook
import win32com.client
import numpy as np
from pathlib import Path     
import os 

def formatName(filename):
    temp = filename.split('\\')
    rawName = temp[len(temp)-1].split('.')
    name = rawName[0]
    return name

def transpose(workbook):
    wb = load_workbook(workbook)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    # Selezionare l'intervallo di celle da trasporre
    cell_range = ws.calculate_dimension()
    range_to_transpose = ws[cell_range]
    transposed_range = np.transpose([[cell.value for cell in row] for row in range_to_transpose])
    for row in transposed_range:
        ws.append(list(row))
    ws.delete_rows(1,max_row)
    wb.save(workbook)

def initWorkbook(wbName):
    return Workbook(wbName)


def createFile(workbook,fileList):
    initWorkbook("")
    wb = load_workbook(workbook)
    ws = wb.active
    col_idx = 1
    for file in fileList:
        pass



def convertToExcel(csvFile,envInfo):
    module_path = Path(f"{os.getcwd()}\modules\Modulo1.bas")
    fileName = formatName(csvFile)
    outfile = f"{envInfo['excelFolder']}{fileName}.xlsx"
    excel = win32com.client.Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(csvFile)
    wb.VBProject.VBComponents.Import(module_path)
    excel.Application.Run("FormatCSV")
    wb.SaveAs(outfile,51)
    excel.Quit()
    transpose(outfile)

    
def copyToFile(workbook,filename):
    outfile = Workbook() # Creo un nuovo file in cui andare a salvare i dati
    outWS = outfile.active
    
    wb = load_workbook(workbook)
    ws = wb.active
    cell_range = ws.calculate_dimension()
    range_to_copy = ws[cell_range]
    cell_list = []
    for row in range_to_copy:
        for cell in row:
            cell_list.append(cell)
    
    for cell in cell_list:
        col = cell.column_letter
        row = cell.row
        outWS[f"{col}{row}"] = cell.value
        outfile.save(filename)
